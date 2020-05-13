# Jeremy Fisher 5/13/2020

try:                                                                    # To stop any errors
    from PIL import Image                                               # I will use this to modify the images.
    import openpyxl, requests, random, urllib.request, os, datetime     # Some more imports I'll need
    from openpyxl import Workbook                                       # Import the workbook
except:                                                                 # This will happen if they don't have one of the imports
    print('You are lacking one of the required modules, makes sure you have: PIL, ' # Explains error
          'openpyxl, requests, random, urllib.reqyests, os, and datetime!')         # See above
    exit()                                                                          # I stop the program to stop any crashes


# Validation of the selected object to make sure it's in the public domain.
def ispublicdomain(id):                                                                             # I define a new function with the param as id
    collection_url = 'https://collectionapi.metmuseum.org/public/collection/v1/objects/' + str(id)  # Makes the url with using the id
    try:                                                                                            # Prevent crashes, most likely due to internet issues
        artwork_data = requests.get(collection_url).json()                                          # I get the info in json format
    except:                                                                                         # This is what happens if error
        print("Couldn't reach the server. Try checking your internet and re-run this program")      # Explains to user
        exit()                                                                                      # Stops the program
    if artwork_data['isPublicDomain'] is False:                                                     # If inside the data the key isPublicDomain is false
        return False                                            # The function returns the value False
    if artwork_data['isPublicDomain'] is True:                  # If inside the data the key isPublicDomain is True
        return True                                             # The function returns True


# Below is users input validation. This makes sure it's a number and a "real" number
def digit_and_range_validation(min, max):                       # I define the function with min and max expected values as the params
    while True:                                                 # The function loop
        userinput = input("Enter a number: ")                   # Makes a variable of the user's input with the instructions
        try:                                                    # Prevent an error
            userinput = int(userinput)                          # Tries to make the input a int
        except:                                                 # If the input is not an int
            print("Please use digits and whole numbers!")       # Instructs the user
            continue                                            # Restarts the loop
        if userinput > max:                                     # Checks if the input is above the expected max
            continue                                            # Restarts the loop
        if userinput < min:                                     # Checks if the input is under the expected min
            continue                                            # Restarts the loop
        return userinput                                        # If all the checks pass, it returns the input


# Validation to make sure the user doesn't chose an empty department.
def emptydepartmentcheck(url):                                  # Makes a function with url as the param
    try:                                                        # Prevents crashing if there's no internet
        object_data = requests.get(url).json()                  # Gets the data
    except:                                                     # If failed, do this
        print("Couldn't reach the server. Try checking your internet and re-run this program")  # Gives info to user
        exit()                                                  # Stops the program
    if object_data['total'] == 0:                               # If inside the data in the total key and the value is 0
        return True                                             # Returns True
    else:                                                       # Otherwise
        return False                                            # The data is full of info and returns False

def main():                                                     # This is the main loop
    # Setting up the api
    department_url = 'https://collectionapi.metmuseum.org/public/collection/v1/departments'  # This is the api url
    try:                                                            # Prevents the crash
        departments_data = requests.get(department_url).json()      # I store the data in a variable, using requests
    except:                                                         # If it does fail...
        print("Couldn't reach the server. Try checking your internet and re-run this program")  # Explains failure
        exit()                                                  # The program is stopped

    departments = []                                            # I create an empty string to append to
    # The loop below will get every department
    for entry in departments_data['departments']:               # For every entry (department) in the department data
        departments.append(entry['displayName'])                # Append the value (department name)

    print("Hello! Welcome to Jeremy's Art Sticker Project")                      # Welcome text for user
    print("Select from one of the following departments, enter a number.")       # Instructs users
    print()                                         # Blank space
    for x, department in enumerate(departments):    # I create a for loop to display all departments using enumerate
        print(str(x + 1) + ".", department)         # I add 1 to x becuase it starts at 0, and I concatenate it with department.

    # Loop tells if the department the users selects is empty or not
    while True:                                                                                            # The loop
        userinput = digit_and_range_validation(1, len(departments))                                        # Uses the validation funtion with the min being 1 and the max being the length of the list +1 (Becuase it starts at 1
        collectionurl = 'https://collectionapi.metmuseum.org/public/collection/v1/objects?departmentIds='  # The url
        if emptydepartmentcheck((collectionurl + str(userinput))) is True:                                 # If the department is empty
            print("Unfortunally, at this time the department you have selected has no works of art, chose another one.")    # Explains to the user
            continue                                        # Restarts the loop
        else:                                               # Otherwise the department has data
            break                                           # So the loop can be broke

    objectsurl = collectionurl + str(userinput)             # Combines the default url with the selected id
    try:                                                    # Prevents the crash
        object_data = requests.get(objectsurl).json()       # Finally gets the data for the last time
    except:                                                 # If it fails
        print("Couldn't reach the server. Try checking your internet and re-run this program")  # Explains
        exit()                                              # Stops the program

    total_objects = object_data['total']                    # Gets the number of artworks
    objectid_list = object_data['objectIDs']                # Gets the list of object ids
    print("Getting artwork...")                             # Tells user what's happening
    print("This might take a few moments.")                 # Explains the wait
    # This loop selects a random artwork from the chosen department
    while True:                                             # Starts the loop
        random_number = random.randint(1, total_objects)    # I use this varaible to select a random value from the objectid list!!!
        random_id = objectid_list[random_number]            # Takes the random id from the list and stores it in a variable
        if ispublicdomain(random_id) is False:              # Checks if the object is in the public domain
            continue                                        # If it's not, the loop restarts
        else:                                               # Otherwise if it's in the public domain
            # Under this I will show the image and see if the user wants it.
            potentartwork_url = 'https://collectionapi.metmuseum.org/public/collection/v1/objects/' + str(random_id)    # The url of the artwork the user might use
            try:                                                                # Checks internet
                potentartwork_data = requests.get(potentartwork_url).json()     # Gets the artwork data
            except:                                                             # If it fails
                print("Couldn't reach the server. Try checking your internet and re-run this program")  # Explains failure
                exit()                                                              # Stops the program
            primaryImageSmall_url = potentartwork_data['primaryImageSmall']         # Gets the thumbnail, easy download, version of the image
            urllib.request.urlretrieve(primaryImageSmall_url, "TempImage.jpg")      # Saves the image as TempImage.jpg
            thumbnail = Image.open('TempImage.jpg')                                 # opens the thumbnail
            thumbnail.show()                                                        # Shows it to the user
            userinput = input('Do you want this image? (y/n): ').lower()            # Asks if the user wants it and lowercases it
            if userinput in ['y', 'yes']:                                           # If the input is the following
                os.remove('TempImage.jpg')                                          # Removes the temp thumbnail
                break                                                               # Breaks the loop
            else:                                                                   # Others wise the answer is no, so
                continue                                                            # Restart the process.

    primaryImage_url = potentartwork_data['primaryImage']                           # Now gets the fullsize image data
    try:                                                                            # Tries to access the new image
        urllib.request.urlretrieve(primaryImage_url, "chosen_artwork_image.jpg")    # Downloads the image
    except:                                                                         # If it fails
        print("Couldn't reach the server. Try checking your internet and re-run this program")  # Tells user why
        exit()                                                                                  # Stops
    image = Image.open('chosen_artwork_image.jpg')                  # Opens the image object
    try:                                                            # Tries to get the folder contents
        stickers = os.listdir('Stickers')                           # IMPORTANT: Makes a list of the stickers so it doesn't matter which stickers are added!!!
    except:                                                         # if it failed
        print("You are missing the 'sticker' folder. Make sure to download it, "               # Explained why it failed
              "and not change the name. Keep it in the same folder as the python program.")    # Explained why it failed
        exit()                                                      # Stops the program
    print()                                                         # Blank space
    print("Chose a meme sticker or select random")                  # Instructs user

    for x, sticker in enumerate(stickers):                          # Lists the sticker options
        print(str(x + 1) + ".", sticker[:-4])                       # Formats the sting that will repeat
    print(str(len(stickers) + 1) + ". Random")                      # Adds the random option

    userinput = digit_and_range_validation(1, (len(stickers) + 1))  # Gets the user input and validates it.

    if userinput == (len(stickers) + 1):                            # if the user chooses the last option (Random)
        userinput = random.randint(1, len(stickers))                # makes the userinput a random number, that represents a sticker

    selected_sticker = stickers[userinput - 1]                      # Finalizes the sticker in a variable.

    try:                                                            # It tries to open the image
        sticker = Image.open('Stickers\\' + selected_sticker, )     # Tries to open
    except:
        print("Couldn't find the Stickers folder. Make sure to download it, and not change the name. Keep it in the same "
              "folder as the python program.")                      # Explains to user
        exit()                                                      # Stops the program

    # I will probaby need to resize the the sticker to thumbnail size
    width, height = image.width, image.height                       # Gets the artwork image height and width
    sticker.resize((round(width * .125), round(height * .125)))     # I resize the sticker to 1/8 the artwork image to keep aspect ratio
    sticker = sticker.rotate(random.randint(-90, 90))             # this rotates the image, but only to 90 either way to it doesn't look odd

    #### THE ISSUE IM HAVING IS THE ROTATION SOMETIMES GOES OUT SIDE THE BOARDER!!!!
    # I have tried many attempts to fix it, in almost all cases the sticker is just fine, but it seems to sometimes not work
    x_axis = random.randint(round(width * .125), width - round((width * .125)) - round(height * .125) -100)     # Gets a random cooridate
    y_axis = random.randint(round(height * .125), height - round((height * .125)) - round(width * .125) -100)   # Gets a random cooridate

    image.paste(sticker.convert('RGBA'), (x_axis, y_axis), sticker.convert('RGBA'))     # Applies the sticker without background
    image.show()                                                                        # Shows the user
    image.save("Art_and_Sticker.jpeg")                              # Saves the final picture
    print()                                                         # Blank space

    try:                                                            # Tries to open a pre-existing exel sheet
        workbook = openpyxl.load_workbook('StickerArtSheet.xlsx')   # Opens the workbook
        workbook.close()                                            # Closes the book if so
    except:                                                         # Otherwise the book doesn't exist, so time to set it up!
        workbook = Workbook()                                   # Calls the work book function
        sheet = workbook.active                                 # Make the sheet active
        sheet.cell(1, 1, 'Title')                               # Makes the first cell the title section
        sheet.cell(1, 2, "Artist")                              # Makes the next section Artist
        sheet.cell(1, 3, "URL")                                 # Makes the next section URl
        sheet.cell(1, 4, 'Sticker')                             # Makes the next section Sticker
        sheet.cell(1, 5, 'Date & Time')                         # Makes the next section Date and time
        workbook.save('StickerArtSheet.xlsx')                   # Saves the excel sheet
        workbook.close()                                        # Closes the sheet

    artbook = openpyxl.load_workbook('StickerArtSheet.xlsx')    # Opens the sheet
    artsheet = artbook.active                                   # Makes it active

    artist = potentartwork_data['artistDisplayName']            # Gets the artist's name
    if artist == '':                                            # Checks if the artist is nothing
        artist = "Unknown or N/A"                               # if it is, makes the artist Unknown

    workbook_data = [str(potentartwork_data['title']), artist, str(primaryImage_url),   # Makes the set of data
                     str(selected_sticker), str(datetime.datetime.today())]             # Indented for readiblity

    columns_list = list(artsheet.columns)               # Gets the columns as a list
    columns1 = columns_list[0]                          # Makes column 1 the first entry of the list
    row = len(columns1) + 1                             # makes row number the one after the prexisteng row

    for thing in workbook_data:                         # Prints to user the info saved
        print(thing)                                    # The actually printing action

    # Due to the loop not working, I hard coded this part
    artsheet.cell(row, 1, potentartwork_data['title'])  # Saves the artwork name
    artsheet.cell(row, 2, artist)                       # Saves the artist
    artsheet.cell(row, 3, primaryImage_url)             # Saves the URL
    artsheet.cell(row, 4, selected_sticker)             # Saves the sticker
    artsheet.cell(row, 5, datetime.datetime.today())    # Saves the time and date
    artbook.save('StickerArtSheet.xlsx')                # Saves the new edits to the workbook


main()                                                  # Runs the main code
